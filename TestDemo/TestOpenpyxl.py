from openpyxl import load_workbook

wb = load_workbook('productsss.xlsx')
print (wb.sheetnames)
ws = wb[wb.sheetnames[0]] # dùng tên đó để lấy sheet đầu tiên và gán cho biến ws.
for row in ws.values:
    for value in row:
        print(value,"\t",end='')
    print("")